import sys
sys.path.append('/home/neuro_sim2/tvb/better_tesi')

import xlrd
from openpyxl import load_workbook
from connectivity import load_mousebrain
import numpy as np


# finding the structures present in Oh et al.
book = xlrd.open_workbook('oh_table1.xls')
sh = book.sheet_by_index(1)
structures_Oh = []

for value in sh.col_values(3):
    if isinstance(value, str):
        structures_Oh.append(value)

structures_Oh.pop(0)
#print(f"Structures in Oh et al.: {structures_Oh}")
print('Len Oh structures: ',len(structures_Oh))


# finding the strucures present in Gozzi et al.
wb = load_workbook(filename='atlas_gozzi.xlsx')
sheet = wb.active
structures_Gozzi = []

for value in range(1, sheet.max_row):
    structure = sheet.cell(row=value, column=3)
    structures_Gozzi.append(structure.value)

structures_Gozzi.pop(0)
#print(f"Structures in Gozzi et al.: {structures_Gozzi}")
print('Len Gozzi structures: ',len(structures_Gozzi))



# compare the two lists
common_regions = []
only_OH = []
only_Gozzi = []

for s in structures_Gozzi:
    for r in structures_Oh:
        if s == r:
            common_regions.append(s)


print('Len common regions', len(common_regions))

only_OH = list(set(structures_Oh)-set(structures_Gozzi))
print('Len only OH', len(only_OH))
only_Gozzi = list(set(structures_Gozzi)-set(structures_Oh))
print('Len only Gozzi: ', len(only_Gozzi))
        

#print(f"Common regions between the two: {common_regions} \nFor a total of {len(common_regions)} common regions")
np.savetxt('common_regs.txt',common_regions,fmt='%s')
np.savetxt('only_oh.txt',only_OH,fmt='%s')
np.savetxt('only_gozzi.txt',only_Gozzi,fmt='%s')

# finding the indexes of the common regions
brain = load_mousebrain("Connectivity_596.h5", norm="log", scale="region")
nreg = len(brain.weights)

common_ids = [i for i in range(nreg) if any(b in brain.region_labels[i] for b in common_regions)]
#print(common_ids)


